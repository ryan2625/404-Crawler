from bs4 import BeautifulSoup
from all_pages import arr
import requests
from openpyxl import Workbook
import time

# Array of links that we have already gotten a 200 status code from 
# (Don't need to recheck these links, save resources)
# Need to update docs
allSafeLinks = []
allBrokenLinks = []
allBrokenLinksTotal = []
# Dictionary to remember status codes of URLs we have already visited
linkCodeMatch = {}

arrTuple =[]
thrownErr = []
# Use to only check first link in our array
n = 0
# Entire page we want to scan
for link in arr:
    n+=1 
    print("NUMBER", n)
    url = link
    tries = 0
    active = ""
    while active == "":
        if (tries >= 2) :
            break
        try: 
            active = requests.get(url, timeout=5) 
        except:
            print("Main page could not be handled, skipping for now")
            time.sleep(8)
            tries+=1
    if (tries >= 2) :
        thrownErr.append(link)
        continue
    doc = BeautifulSoup(active.content, "html.parser")
    all_links = doc.find_all('a', href=True)
    for subLink in all_links:
        linkHref = subLink['href']
        # Check if link doesn't include https (I.E href = "/myirem")
        if (len(linkHref) > 0) and linkHref[0] == "/":
            temp = linkHref
            linkHref = "https://www.irem.org" + temp
        
        
        # If link is already returning 200, don't need to check again
        if (linkHref not in allSafeLinks and linkHref not in allBrokenLinks):
            if len(linkHref) > 3 and linkHref[:4] == "http" and linkHref != "https://":
                print(linkHref)
                try: 
                    code = requests.get(linkHref, timeout=30)
                except Exception as e:
                    allBrokenLinks.append(linkHref)
                    allBrokenLinksTotal.append(linkHref)
                    arrTuple.append((link, linkHref, "ERROUT"))
                    print("ERR", e)
                    continue
                code = str(code.status_code) 
                linkCodeMatch[linkHref] = code
            # Only mark link as safe if we get 200 level response
                if code[0] == "2": 
                    allSafeLinks.append(linkHref)
                else:
                    arrTuple.append((link, linkHref, code))
                    allBrokenLinks.append(linkHref)
            # If link is empty, etc    
            else:
                allBrokenLinksTotal.append(linkHref)
        
        else:
            if (linkHref in allBrokenLinks):
                if (linkHref not in allBrokenLinksTotal):
                    if (linkHref != "https://twitter.com/IREM_info" or linkHref != "https://www.linkedin.com/company/institute-of-real-estate-management"):
                        code = linkCodeMatch[linkHref]
                        arrTuple.append((link, linkHref, code))
            

    if (n % 50 ==0):
        wb = Workbook()
        ws = wb.active 
        for tuples in arrTuple:
            ws.append([tuples[0], tuples[1], tuples[2]])
        wb.save(str(n) + ".xlsx")

wb = Workbook()
ws = wb.active
for tuples in arrTuple:
    ws.append([tuples[0], tuples[1], tuples[2]])
wb.save("all_broken.xlsx")

wb = Workbook()
ws = wb.active
for ele in allBrokenLinksTotal:
    ws.append([ele])
wb.save("misc_broken.xlsx") 