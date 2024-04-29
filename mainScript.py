import time
from bs4 import BeautifulSoup
from openpyxl import Workbook
from all_pages import arr
from all_safe import safeArr
from handleLogin import getSession
from bad_tuple import bad
'''
TODO:  
MORE testing - create excel sheet for 5 links in arr2
For each of the sheets manually check and make sure its capturing all the links you want

OR just create your own index html and manually input links
'''

LOGIN_PAGE = "https://www.irem.org/sso/login.aspx"
allSafeLinks = safeArr
allBrokenLinks = []
miscBroken = []
linkCodeMatch = {}
arrTuple = bad
session = getSession(LOGIN_PAGE)
n = 0

def handleNewLink(linkHref, link):
    global allBrokenLinks
    global arrTuple
    global allSafeLinks
    global session
    global linkCodeMatch
    if len(linkHref) > 3 and (linkHref[:4] == "http" or linkHref[:3] == "www"):
        print(linkHref) 
        try: 
            code = session.get(linkHref, timeout=30)
        except Exception as e:
            linkCodeMatch[linkHref] = "ERR: TIMEOUT"
            allBrokenLinks.append(linkHref)
            arrTuple.append((link, linkHref, f"ERR TIMEOUT: {e}"))
            print("ERR", e)
            return
        code = str(code.status_code) 
    # Only mark link as safe if we get 200 level response
        if code[0] == "2": 
            allSafeLinks.append(linkHref)
        elif code == "404":
            linkCodeMatch[linkHref] = code
            arrTuple.append((link, linkHref, code))
            allBrokenLinks.append(linkHref)
        else:
            linkCodeMatch[linkHref] = code
            allBrokenLinks.append(linkHref)
            miscBroken.append((link, linkHref))

# If link is empty, href="javascript:void(0)"", #, etc
    else:
        miscBroken.append((link, linkHref))

def handleSubLinks(all_links, link):
    global arrTuple
    for subLink in all_links:
        linkHref = subLink['href'].strip()
        # Check if link doesn't include https (I.E href = "/myirem")
        if (len(linkHref) > 0) and linkHref[0] == "/":
            temp = linkHref
            linkHref = "https://www.irem.org" + temp
        
        # Linkedin and twitter don't like bots
        if len(linkHref) >=20 and (linkHref[:20] == "https://twitter.com/" or linkHref[:20] == "https://www.linkedin"):
            continue

        # If we already checked this link, no need to check again
        if (linkHref not in allSafeLinks and linkHref not in allBrokenLinks):
            handleNewLink(linkHref, link)
        
        else:
            if (linkHref in allBrokenLinks):
                code = linkCodeMatch[linkHref]
                if (code == "404"):
                    arrTuple.append((link, linkHref, code))

def saveToExcel():
    wb = Workbook()
    ws = wb.active
    for tuples in arrTuple:
        ws.append([tuples[0], tuples[1], tuples[2]])
    wb.save("all_broken.xlsx")

    wb = Workbook()
    ws = wb.active
    for ele in miscBroken:
        ws.append([ele[0], ele[1]])
    wb.save("misc_broken.xlsx") 

    wb = Workbook()
    ws = wb.active 
    for ele in allSafeLinks:
        ws.append([ele])
    wb.save("all_safe.xlsx")

def main():
    global n
    global allBrokenLinks
    global arrTuple
    global allSafeLinks
    global session
    for link in arr:
        n+=1 
        print("LINK ", n)
        url = link
        tries = 0
        active = ""
        err = ""
        while active == "":
            if (tries >= 2) :
                break
            try: 
                active = session.get(url, timeout=30) 
            except Exception as e:
                err=e
                print(f"Main page timed out...{e}")
                time.sleep(5)
                tries+=1
        if (tries >= 2) :
            linkCodeMatch[link] = "ERR: TIMEOUT"
            allBrokenLinks.append(link)
            arrTuple.append((link, link, f"ERR MAIN TIMEOUT: {err}"))
            continue
        doc = BeautifulSoup(active.content, "html.parser")
        all_links = doc.find_all('a', href=True)
        handleSubLinks(all_links, link)
    saveToExcel()

main()